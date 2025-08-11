# pip install openpyxl
import re
from openpyxl import load_workbook
from pathlib import Path

EXCEL_PATH = "test.xlsx"

ID_HEADER = "Condition Number"
NAME_HEADER = "Content Classifier Name"
TYPE_HEADER = "Type"
THRESH_HEADER = "Threshold"

def normalize_id(x):
    if x is None:
        return ""
    s = str(x).strip()
    try:
        return str(int(float(s)))  # e.g., 1.0 -> "1"
    except:
        return s

def escape_md_cell(text):
    return str(text).replace("|", r"\|")

def already_wrapped(s: str) -> bool:
    s = s.strip()
    return s.startswith("(") and s.endswith(")")

def sanitize_filename(name: str) -> str:
    # Replace characters not allowed on Windows/macOS/Linux filenames
    sanitized = re.sub(r'[\\/:*?"<>|]', "_", str(name).strip())
    # Avoid trailing dots/spaces (Windows)
    sanitized = sanitized.rstrip(" .")
    return sanitized or "sheet"

def process_sheet(ws):
    # 1) Read original boolean logic from A1
    original_logic = str(ws["A1"].value or "")

    # 2) Find the header columns (row 2)
    headers = { str(ws.cell(row=2, column=c).value or "").strip(): c
                for c in range(1, ws.max_column + 1) }

    for h in (ID_HEADER, NAME_HEADER, TYPE_HEADER, THRESH_HEADER):
        if h not in headers:
            raise KeyError(f"Missing required header '{h}' in row 2. Found: {list(headers.keys())}")

    id_col = headers[ID_HEADER]
    name_col = headers[NAME_HEADER]
    type_col = headers[TYPE_HEADER]
    thresh_col = headers[THRESH_HEADER]

    # 3) Build mapping and collect table rows
    mapping = {}  # "1" -> "ClassifierName"
    table_rows = []
    for r in range(3, ws.max_row + 1):
        id_val = ws.cell(row=r, column=id_col).value
        name_val = ws.cell(row=r, column=name_col).value
        type_val = ws.cell(row=r, column=type_col).value
        thresh_val = ws.cell(row=r, column=thresh_col).value
        if id_val is None or name_val is None:
            continue
        key = normalize_id(id_val)
        if key:
            mapping[key] = str(name_val).strip()
        table_rows.append([str(id_val or ""), str(name_val or ""), str(type_val or ""), str(thresh_val or "")])

    # 4) Replace whole-number tokens with classifier names
    expanded_logic = re.sub(r"\b\d+\b", lambda m: mapping.get(m.group(0), m.group(0)), original_logic)

    # 5) Format:
    #    - split by OR into blocks (each becomes its own line)
    #    - inside each block, keep AND items on the same line
    #    - if an AND block has > 2 items, wrap the whole line in parentheses
    or_blocks = re.split(r"\bOR\b", expanded_logic, flags=re.IGNORECASE)

    formatted_blocks = []
    for block in or_blocks:
        block = re.sub(r"\s+", " ", block.strip())  # normalize spaces
        if not block:
            continue

        and_terms = [re.sub(r"\s+", " ", t.strip()) for t in re.split(r"\bAND\b", block, flags=re.IGNORECASE)]
        and_terms = [t for t in and_terms if t]

        if len(and_terms) == 1:
            line = and_terms[0]
        else:
            line = " AND ".join(and_terms)
            if len(and_terms) > 2 and not already_wrapped(line):
                line = f"({line})"

        formatted_blocks.append(line)

    # Build final Expanded Boolean text:
    # - Each OR is on its own indented line ("  OR")
    # - Every line (including OR lines) ends with two spaces for Markdown line breaks
    output_lines = []
    for i, line in enumerate(formatted_blocks):
        if i > 0:
            output_lines.append("  OR  ")   # leading indent + trailing two spaces
        output_lines.append(f"{line}  ")    # trailing two spaces

    expanded_logic_formatted = "\n".join(output_lines).rstrip()

    # 6) Build Markdown output for this sheet
    md_lines = []
    md_lines.append("# DLP Classifier Logic Expansion\n")
    md_lines.append("## Input Classifiers\n")
    md_lines.append("| Condition Number | Content Classifier Name | Type | Threshold |")
    md_lines.append("| --- | --- | --- | --- |")
    for row in table_rows:
        md_lines.append("| " + " | ".join(escape_md_cell(v) for v in row) + " |")

    md_lines.append("\n## Original Boolean Logic\n")
    md_lines.append("```")
    md_lines.append(original_logic)
    md_lines.append("```")

    md_lines.append("\n## Expanded Boolean Logic (OR on its own indented line)\n")
    md_lines.append(expanded_logic_formatted)

    return "\n".join(md_lines)

# -------- Main: iterate all sheets and write <sheetname>.md --------
wb = load_workbook(EXCEL_PATH)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    md_content = process_sheet(ws)
    out_name = f"{sanitize_filename(sheet_name)}.md"
    Path(out_name).write_text(md_content, encoding="utf-8")
    print(f"Wrote {out_name}")
